import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
import os
import platform
import configparser

def leer_config():
    config = configparser.ConfigParser()
    config.read("config.ini")
    return config["mysql"]

def conectar_bd():
    try:
        config = leer_config()
        return mysql.connector.connect(
            host=config.get("host", "localhost"),
            user=config.get("user", "root"),
            password=config.get("password", ""),
            database=config.get("database", "gimnas")
        )
    except mysql.connector.Error as err:
        messagebox.showerror("Error", f"No se pudo conectar a la base de datos: {err}")
        return None

def generar_excel(tipo, datos, archivo):
    wb = Workbook()
    ws = wb.active
    titulo = "esporadics Actius" if tipo == "actius" else "esporadics Inactius"
    ws.title = titulo

    encabezados = [
        'ID', 'DNI', 'Nom', 'Carrer', 'Codipostal', 'Poblacio', 'Provincia', 'email', 
        'Data_naixement', 'Telefon1', 'Telefon2', 'Telefon3', 'Numero_Conta', 'Sepa', 
        'Activitats', 'Quantitat', 'Alta', 'Baixa', 'Facial', 'Data_Inici_activitat', 
        'Usuari', 'Descompte', 'Total', 'Temps_descompte', 'Extres', 'En_ma'
    ]
    ws.append(encabezados)

    for dato in datos:
        ws.append([dato.get(campo, "") for campo in encabezados])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].auto_size = True

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border

    wb.save(archivo)

    try:
        if platform.system() == "Windows":
            os.startfile(archivo)
        elif platform.system() == "Darwin":
            os.system(f"open '{archivo}'")
        else:
            os.system(f"xdg-open '{archivo}'")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir el archivo automáticamente: {e}")

def obtener_datos(tipo):
    conexion = conectar_bd()
    if not conexion:
        return []
    cursor = conexion.cursor(dictionary=True)
    query = (
        "SELECT * FROM esporadics WHERE Activitats IS NOT NULL AND Activitats != ''"
        if tipo == "actius"
        else "SELECT * FROM esporadics WHERE Activitats IS NULL OR Activitats = ''"
    )
    cursor.execute(query)
    datos = cursor.fetchall()
    conexion.close()
    return datos

def generar_archivo():
    tipo = combo_tipo.get()
    if not tipo:
        messagebox.showwarning("Atención", "Seleccione un tipo de lista.")
        return

    archivo = filedialog.asksaveasfilename(
        defaultextension=".xlsx", 
        filetypes=[("Excel files", "*.xlsx")],
        title="Guardar archivo Excel"
    )
    if not archivo:
        return

    datos = obtener_datos(tipo)
    if not datos:
        messagebox.showinfo("Información", "No se encontraron datos para exportar.")
        return

    generar_excel(tipo, datos, archivo)
    messagebox.showinfo("Éxito", f"Archivo guardado correctamente en {archivo}")

root = tk.Tk()
root.title("Generar Llistes de esporadics")
root.geometry("400x200")

tk.Label(root, text="Seleccioneu el tipus de llista:", font=("Arial", 12)).pack(pady=10)
combo_tipo = ttk.Combobox(root, state="readonly", values=["actius", "inactius"])
combo_tipo.pack(pady=10)

btn_generar = tk.Button(root, text="Generar Excel", command=generar_archivo, font=("Arial", 12), bg="blue", fg="white")
btn_generar.pack(pady=20)

root.mainloop()
